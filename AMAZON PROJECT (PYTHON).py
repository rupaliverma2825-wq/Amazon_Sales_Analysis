import pandas as pd

a=pd.read_csv("Amazon.csv",encoding="utf-8")
print(a)

print(a.columns)

print(a.duplicated())
print(a.info())

# changing datatype

a['OrderDate']= pd.to_datetime(a["OrderDate"])
print(a.info())

a=a.drop(columns=['Year'])
print(a)

a['year']=a['OrderDate'].dt.year
print(a)

print(a.info())

pd.set_option("display.max_columns",None)
print(a.info())

a.to_csv("Amazon_cleaned.csv")
print(a)

import pandas as pd 
from sqlalchemy import create_engine

engine = create_engine (
    "postgresql+psycopg2://postgres:Rupali0055@localhost:5432/Amazon"
)  
print("successfully connected")

top_5_selling_products=pd.read_sql(
    """SELECT
	PRODUCT_NAME,
	SUM(TOTAL_AMOUNT) AS TOTAL
FROM
	AMAZON
GROUP BY
	PRODUCT_NAME,
	TOTAL_AMOUNT
ORDER BY
	TOTAL_AMOUNT DESC
LIMIT
	5;
""",engine
)

print(top_5_selling_products)

import matplotlib.pyplot as plt

plt.bar(top_5_selling_products["product_name"],top_5_selling_products["total"])
plt.title("TOP 5 SELLING PRODUCTS")
plt.xticks(rotation=45,ha="right")
plt.tight_layout()
plt.savefig("top_5_selling_products.png")
plt.show()


total_sales_by_Category= pd.read_sql(
	"""SELECT
	CATEGORY,
	SUM(TOTAL_AMOUNT) AS TOTAL_AMOUNT
FROM
	AMAZON
GROUP BY
	CATEGORY
ORDER BY
	TOTAL_AMOUNT DESC;""",engine
)

print(total_sales_by_Category)

plt.bar(total_sales_by_Category["category"],total_sales_by_Category["total_amount"])
plt.title("TOTAL SALES AMOUNT BY CATEGORY")
plt.xticks(rotation=45,ha="right")
plt.tight_layout()
plt.savefig("total_sales_by_Category.png")
plt.show()


top_5_brands_by_total_sales_amount= pd.read_sql(
	"""SELECT
	BRAND,
	SUM(TOTAL_AMOUNT) AS TOTAL_AMOUNT
FROM
	AMAZON
GROUP BY
	BRAND
ORDER BY
	TOTAL_AMOUNT DESC
LIMIT
	5;""",engine
)

print(top_5_brands_by_total_sales_amount)

plt.bar(top_5_brands_by_total_sales_amount["brand"],top_5_brands_by_total_sales_amount["total_amount"])
plt.title("TOP 5 BRANDS BY TOTAL SALES AMOUNT")
plt.xticks(rotation=45,ha="right")
plt.tight_layout()
plt.savefig("top_5_brands_by_total_sales_amount.png")
plt.show()

top5_customers_with_more_than_2_orders_in_2024= pd.read_sql(
	"""SELECT
	CUSTOMER_ID,
	CUSTOMER_NAME,
	COUNT(*) AS ORDER_COUNT
FROM
	AMAZON
WHERE
	YEAR = 2024
GROUP BY
	CUSTOMER_ID,
	CUSTOMER_NAME
HAVING
	COUNT(*) > 2
ORDER BY
	ORDER_COUNT DESC
limit 5;
""",engine
)

print(top5_customers_with_more_than_2_orders_in_2024)
plt.bar(top5_customers_with_more_than_2_orders_in_2024["customer_name"],
        top5_customers_with_more_than_2_orders_in_2024["order_count"])
plt.title("TOP 5 CUSTOMERS WITH MORE THAN 2 ORDERS IN 2024")
plt.xticks(rotation=45,ha="right")
plt.tight_layout()
plt.savefig("top5_customers_with_more_than_2_orders_in_2024.png")
plt.show()

top_5_products_by_total_sales_amount_in_2024=pd.read_sql(
	"""SELECT
	PRODUCT_NAME,
	SUM(TOTAL_AMOUNT) AS TOTAL_AMOUNT
FROM
	AMAZON
WHERE
	YEAR = 2024
GROUP BY
	PRODUCT_NAME
ORDER BY
	TOTAL_AMOUNT DESC
LIMIT
	5;""",engine
)
print(top_5_products_by_total_sales_amount_in_2024)

plt.bar(top_5_products_by_total_sales_amount_in_2024["product_name"],
        top_5_products_by_total_sales_amount_in_2024["total_amount"])
plt.title("TOP 5 PRODUCTS BY TOTAL SALES AMOUNT IN 2024")
plt.xticks(rotation=45,ha="right")
plt.tight_layout()
plt.savefig("top_5_products_by_total_sales_amount_in_2024.png")
plt.show()

top_5_products_by_total_sales_amount_in_2023= pd.read_sql(
	"""SELECT
	PRODUCT_NAME,
	SUM(TOTAL_AMOUNT) AS TOTAL_AMOUNT
FROM
	AMAZON
WHERE
	YEAR = 2023
GROUP BY
	PRODUCT_NAME
ORDER BY
	TOTAL_AMOUNT DESC
LIMIT
	5;""",engine
)
print(top_5_products_by_total_sales_amount_in_2023)

plt.bar(top_5_products_by_total_sales_amount_in_2023["product_name"],
        top_5_products_by_total_sales_amount_in_2023["total_amount"])
plt.title("TOP 5 PRODUCTS BY TOTAL SALES AMOUNT IN 2023")
plt.xticks(rotation=45,ha="right")
plt.tight_layout()
plt.savefig("top_5_products_by_total_sales_amount_in_2023")
plt.show()

max_price_of_product= pd.read_sql(
	"""SELECT
	PRODUCT_NAME,
	TOTAL_AMOUNT
FROM
	AMAZON
ORDER BY 
   TOTAL_AMOUNT DESC
LIMIT 1;""",engine
)
print(max_price_of_product)

plt.bar(max_price_of_product["product_name"],max_price_of_product["total_amount"])
plt.title("MAXIMUM PRICE OF PRODUCT")
plt.tight_layout()
plt.savefig("max_price_of_product")
plt.show()

total_discount_by_categories= pd.read_sql(
	"""SELECT
	CATEGORY,
	SUM(DISCOUNT) AS TOTAL_DISCOUNT
FROM
	AMAZON
GROUP BY
	CATEGORY
ORDER BY
	TOTAL_DISCOUNT DESC;""",engine
)

print(total_discount_by_categories)

plt.bar(total_discount_by_categories["category"],total_discount_by_categories["total_discount"])
plt.title("TOTAL DISCOUNT AMOUNT BY CATEGORY")
plt.xticks(rotation=45,ha="right")
plt.tight_layout()
plt.savefig("total_discount_by_categories")
plt.show()

highest_tax_paid_by_categories= pd.read_sql(
	"""SELECT
	CATEGORY,
	SUM(TAX):: NUMERIC AS TOTAL_TAX
FROM
	AMAZON
GROUP BY
	CATEGORY
ORDER BY
	TOTAL_TAX DESC;
""",engine
)

print(highest_tax_paid_by_categories)

plt.bar(highest_tax_paid_by_categories["category"],highest_tax_paid_by_categories["total_tax"])
plt.title("HIGHEST TAX PAID BY CATEGORIES")
plt.xticks(rotation=45,ha="right")
plt.tight_layout()
plt.savefig("highest_tax_paid_by_categories")
plt.show()

top_5_prod_with_highest_shipping=pd.read_sql(
	"""SELECT
	PRODUCT_NAME,
	SUM(SHIPPING_COST) AS SHIPPING_COST
FROM
	AMAZON
GROUP BY
	PRODUCT_NAME
ORDER BY
	SHIPPING_COST DESC
LIMIT
	5;""",engine
)

print(top_5_prod_with_highest_shipping)

plt.bar(top_5_prod_with_highest_shipping["product_name"],top_5_prod_with_highest_shipping["shipping_cost"])
plt.title("TOP 5 PRODUCTS WITH HIGHEST SHIPPING COST")
plt.xticks(rotation=45,ha="right")
plt.tight_layout()
plt.savefig("product_with_highest_shipping_costs")
plt.show()

most_preferable_payment_method_by_cx=pd.read_sql(
	"""SELECT
	PAYMENT_METHOD,
	SUM(TOTAL_AMOUNT) AS TOTAL_AMOUNT
FROM
	AMAZON
GROUP BY
	PAYMENT_METHOD
ORDER BY
	TOTAL_AMOUNT DESC;""",engine
)

print(most_preferable_payment_method_by_cx)

plt.bar(most_preferable_payment_method_by_cx["payment_method"],
        most_preferable_payment_method_by_cx["total_amount"])

plt.title("MOST PREFERABLE PAYMENT METHOD BY CUSTOMERS")
plt.xticks(rotation=45,ha="right")
plt.tight_layout()
plt.savefig("most_preferable_payment_method_by_cx")
plt.show()

total_sales_by_country=pd.read_sql(
	"""SELECT
	COUNTRY,
    SUM(TOTAL_AMOUNT) AS TOTAL_AMOUNT
FROM
	AMAZON
GROUP BY
	COUNTRY
ORDER BY
	TOTAL_AMOUNT DESC;
""",engine
)

pd.options.display.float_format = '{:,.2f}'.format

print(total_sales_by_country)

plt.bar(total_sales_by_country["country"],total_sales_by_country["total_amount"])
plt.title("TOTAL SALES AMOUNT BY COUNTRY")
plt.xticks(rotation=45,ha="right")
plt.tight_layout()
plt.savefig("total_sales_by_country")
plt.show()

order_status_breakdown=pd.read_sql(
	"""SELECT ORDER_STATUS, COUNT (*)AS TOTAL
FROM AMAZON
GROUP BY ORDER_STATUS 
ORDER BY TOTAL DESC;""",engine
)

print(order_status_breakdown)

plt.bar(order_status_breakdown["order_status"],order_status_breakdown["total"])
plt.title("ORDER STATUS BREAKDOWN")
plt.xticks(rotation=45,ha="right")
plt.tight_layout()
plt.savefig("order_status_breakdown")
plt.show()

category_percentage= a.groupby('Category')['TotalAmount'].sum()/a['TotalAmount'].sum()*100
category_percentage=category_percentage.reset_index().rename(columns={'TotalAmount':'Percentage'})
category_percentage['Percentage']=category_percentage['Percentage'].round(2)
category_percentage=category_percentage.sort_values(by='Percentage',ascending=False)
print(category_percentage)

plt.bar(category_percentage["Category"],category_percentage["Percentage"])
plt.title("CATEGORY CONTRIBUTION TO TOTAL SALES")
plt.xticks(rotation=45,ha="right")
plt.tight_layout()
plt.savefig("category_percentage.jpg")
plt.show()


country_per=a.groupby("Country")["TotalAmount"].sum()/a["TotalAmount"].sum()*100
country_per=country_per.reset_index().rename(columns={"TotalAmount" : "Percentage"})
country_per['Percentage']=country_per['Percentage'].round(2)
country_per=country_per.sort_values(by="Percentage",ascending= False)
print(country_per)

plt.bar(country_per["Country"],country_per["Percentage"])
plt.title("COUNTRY CONTRIBUTION TO TOTAL SALES")
plt.xticks(rotation=45,ha="right")
plt.tight_layout()
plt.savefig("Country_percentage.jpg")
plt.show()

order_status_per= a.groupby("OrderStatus")["TotalAmount"].sum()/a["TotalAmount"].sum()*100
order_status_per=order_status_per.reset_index().rename(columns={"TotalAmount":"Percentage"})
order_status_per=order_status_per.sort_values(by="Percentage",ascending=False)
print(order_status_per)

plt.bar(order_status_per["OrderStatus"],order_status_per["Percentage"])
plt.title("ORDER STATUS CONTRIBUTION")
plt.xticks(rotation=45,ha="right")
plt.tight_layout()
plt.savefig("order_status_percentage")
plt.show()

# CONVERTING THEM INTO EXCEL FILES 

from openpyxl import Workbook
from openpyxl.drawing.image import Image

def create_excel_sheets (df,chart_path,file_name,insight_text):
    wb=Workbook()
    ws=wb.active
    ws.title="Analysis"

    ws.append(list(df.columns))
    for row in df.itertuples (index= False):   
     ws.append(row)
    
    chart_row=len(df)+3
    ws.add_image(Image(chart_path),f"A{chart_row}")

    insight_row=chart_row+20
    ws[f"A{insight_row}"]= "Insight :"
    ws[f"B{insight_row}"]= insight_text

    wb.save (file_name)
    print(f"{file_name} created")

files_info=[
	[top_5_selling_products,"top_5_selling_products.png","Top_5_selling_producys.xlsx",
  "Puzzle 1000pc is the top-performing product with total sales of ₹10,604.94."],
 
 [total_sales_by_Category,"total_sales_by_Category.png","Total_sales_by_category.xlsx",
  "Electronics is the top-performing category, generating total sales of ₹46,752,651.54."],
 
 [category_percentage,"category_percentage.jpg","Category_percentage.xlsx",
  "Electronics is the top-performing category, contributing 16.97% to total sales."],
 
 [top_5_brands_by_total_sales_amount,"top_5_brands_by_total_sales_amount.png","Top_5_brands.xlsx",
  "CoreTech is the top-performing brand, generating total sales of ₹28,031,161.77."],
 
 [top5_customers_with_more_than_2_orders_in_2024,"top5_customers_with_more_than_2_orders_in_2024.png",
  "top5_customers_with_more_than_2_orders_in_2024.xlsx","Customers such as Aditya Gupta, Aarav Mehta, Aditya Kapoor, Aman Kumar, and Arjun Patel placed more than 2 orders each in 2024, indicating high repeat purchase activity."],
 
 [top_5_products_by_total_sales_amount_in_2024,"top_5_products_by_total_sales_amount_in_2024.png",
  "top_5_products_by_total_sales_amount_in_2024.xlsx","Water Bottle is the top-performing product in 2024, generating total sales of ₹1,195,683.45, followed closely by Dress Shirt, Power Bank 20000mAh, Action Camera, and Wireless Charger."],

 [top_5_products_by_total_sales_amount_in_2023,"top_5_products_by_total_sales_amount_in_2023.png"
  ,"top_5_products_by_total_sales_amount_in_2023.xlsx","Puzzle 1000pc emerged as the top-performing product in 2023, generate total sales of ₹1,262,004.90"],

 [max_price_of_product,"max_price_of_product.png","max_price_of_product.xlsx","Puzzle 1000pc has the highest unit price of ₹3,534.98."],
 
 [total_discount_by_categories,"total_discount_by_categories.png","Highest_discount_by_categories.xlsx","The Electronics category is the top-performing category and has the highest total discount provided"],
 [highest_tax_paid_by_categories,"highest_tax_paid_by_categories.png","highest_tax_paid_by_categories.xlsx","Electronics is the top category in terms of total tax paid i.e ₹3,510,610"],
 [top_5_prod_with_highest_shipping,"product_with_highest_shipping_costs.png","product_with_highest_shipping_cost","Mechanical Keyboard is the product with the highest shipping cost of ₹46,436.23"],
 [most_preferable_payment_method_by_cx,"most_preferable_payment_method_by_cx.png","most_preferable_payment_method_by_cx.xlsx","Credit Card is the most preferred payment method by customers, with total transactions amounting to ₹96,366,476.07."],
 [total_sales_by_country,"total_sales_by_country.png","total_sales_by_country.xlsx","“The United States is the top-selling country, generating total sales of ₹192,930,145.50"],
 [country_per,"Country_percentage.jpg","Country_percentage.xlsx","The United States is the top-selling country, accounting for 70% of total sales"],
 [order_status_breakdown,"order_status_breakdown.png","order_status_breakdown.xlsx","A total of 223,884 orders were delivered, indicating strong fulfillment performance"],
 [order_status_per,"order_status_percentage.png","order_status_percentage.xlsx","Approximately 74.46% of orders were delivered, 15.34% were shipped, 4.07% were pending, 3.10% were cancelled, and 3.03% were returned, indicating strong overall fulfillment performance."]
 
]

for i in files_info:
    create_excel_sheets (* i)